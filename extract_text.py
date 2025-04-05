import fitz  # PyMuPDF
import easyocr
import cv2
import numpy as np
import os
import pandas as pd
from pathlib import Path
import re
from datetime import datetime
from openpyxl import load_workbook

# Initialize the OCR reader
reader = easyocr.Reader(['en', 'fr'])  # Adding French language support


def preprocess_image(img):
    """Apply image preprocessing to improve OCR accuracy"""
    # Convert to grayscale if not already
    if len(img.shape) == 3:
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    else:
        gray = img

    # Apply adaptive threshold
    binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                   cv2.THRESH_BINARY, 11, 2)

    # Noise removal
    denoised = cv2.fastNlMeansDenoising(binary, None, 10, 7, 21)

    return denoised


def extract_text_from_pdf(pdf_path, save_images=True):
    """Extract text from a scanned PDF using OCR"""
    # Create output directories if they don't exist
    output_dir = Path("extracted_images")
    output_dir.mkdir(exist_ok=True)

    print(f"Processing PDF: {pdf_path}")
    doc = fitz.open(pdf_path)
    all_text = []

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)

        # Render the page to an image at higher resolution for better OCR
        pix = page.get_pixmap(dpi=300)
        img_data = pix.samples
        img = np.frombuffer(img_data, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)

        # If grayscale, convert to BGR for consistent processing
        if pix.n == 1:
            img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)

        # Convert to grayscale for OCR processing
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # Preprocess the image
        processed_img = preprocess_image(gray)

        # Save the processed image if requested
        if save_images:
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]
            img_path = f"{base_name}_page{page_num + 1}.png"
            proc_path = f"{base_name}_page{page_num + 1}_processed.png"
            cv2.imwrite(str(output_dir / img_path), img)
            cv2.imwrite(str(output_dir / proc_path), processed_img)

        # Perform OCR on the processed image
        results = reader.readtext(processed_img)

        # Extract and collect text with position information
        page_text = []
        for (bbox, text, confidence) in results:
            if confidence > 0.2:  # Filter very low confidence results
                top_left, top_right, bottom_right, bottom_left = bbox
                x = (top_left[0] + bottom_right[0]) / 2
                y = (top_left[1] + bottom_right[1]) / 2

                print(f"Detected: '{text}' at ({x:.1f}, {y:.1f}) - Confidence: {confidence:.2f}")
                page_text.append({
                    'text': text,
                    'x': x,
                    'y': y,
                    'confidence': confidence,
                    'bbox': bbox,
                    'page': page_num
                })

        all_text.extend(page_text)

    # Close the document when done
    doc.close()
    return all_text


def extract_table_data(text_items):
    """Extract trip details from the declaration table"""
    # Create a dict to store trip information
    trips = []

    # Find table boundaries by looking for header text
    table_headers = ["Date", "Heure", "Pays", "Ville", "Nature de la dépense"]
    header_items = []
    for header in table_headers:
        for item in text_items:
            if header in item['text']:
                header_items.append(item)
                break

    if not header_items:
        print("Could not locate expense table headers")
        return trips

    # Sort header items by x position to get column order
    header_items.sort(key=lambda x: x['x'])

    # Get the y-range where we expect table data (below headers)
    min_y = max([h['y'] for h in header_items]) + 10

    # Find table entries by looking at content in table rows
    for item in text_items:
        # Skip if above or at header level
        if item['y'] <= min_y:
            continue

        # Check if item contains travel data (e.g., dates, flight info)
        date_pattern = r'\d{2}[.-]\d{2}[-.]?\d{0,4}'
        contains_date = re.search(date_pattern, item['text'])
        is_flight = 'flight' in item['text'].lower() or 'vol' in item['text'].lower()
        is_destination = any(
            city in item['text'].lower() for city in ['luxembourg', 'hamburg', 'nancy', 'vienna', 'paris', 'germany'])

        if contains_date or is_flight or is_destination:
            # Look for nearby text to build a complete trip entry
            x, y = item['x'], item['y']
            related_items = [i for i in text_items if abs(i['y'] - y) < 20]  # Items on same row

            # Build trip entry from related items
            trip_entry = {
                'date': '',
                'destination': '',
                'description': '',
                'amount': ''
            }

            for related in related_items:
                text = related['text'].strip()

                # Try to identify what kind of data this is
                if re.search(date_pattern, text):
                    trip_entry['date'] = text
                elif is_destination or any(city in text.lower() for city in
                                           ['luxembourg', 'hamburg', 'nancy', 'vienna', 'paris', 'germany']):
                    trip_entry['destination'] = text
                elif is_flight or 'flight' in text.lower() or 'vol' in text.lower():
                    trip_entry['description'] = text
                elif re.search(r'\d+[.,]\d{2}', text):  # Amounts like 123.45
                    trip_entry['amount'] = text

            # Only add if we have at least date or destination
            if trip_entry['date'] or trip_entry['destination']:
                trips.append(trip_entry)

    # Deduplicate trips
    unique_trips = []
    for trip in trips:
        if trip not in unique_trips:
            unique_trips.append(trip)

    return unique_trips


def extract_person_info(text_items):
    """Extract personal information from the form"""
    person_info = {
        'last_name': '',
        'first_name': '',
        'bank_account': '',
        'travel_dates': ''
    }

    # Look for field labels and extract corresponding values
    for i, item in enumerate(text_items):
        text_lower = item['text'].lower()

        # Last name
        if 'nom' in text_lower and ('famille' in text_lower or 'family' in text_lower):
            # Look for text in the same row or below
            x, y = item['x'], item['y']
            for j in range(i + 1, len(text_items)):
                if (abs(text_items[j]['y'] - y) < 20 or
                        (text_items[j]['y'] > y and text_items[j]['y'] - y < 50)):
                    person_info['last_name'] = text_items[j]['text'].strip()
                    break

        # First name
        if 'prénom' in text_lower or 'first name' in text_lower:
            x, y = item['x'], item['y']
            for j in range(i + 1, len(text_items)):
                if (abs(text_items[j]['y'] - y) < 20 or
                        (text_items[j]['y'] > y and text_items[j]['y'] - y < 50)):
                    person_info['first_name'] = text_items[j]['text'].strip()
                    break

        # Bank account
        if 'compte bancaire' in text_lower or 'bank account' in text_lower:
            x, y = item['x'], item['y']
            for j in range(i + 1, len(text_items)):
                if (abs(text_items[j]['y'] - y) < 20 or
                        (text_items[j]['y'] > y and text_items[j]['y'] - y < 50)):
                    # Format looks like IBAN
                    if re.search(r'[A-Z]{2}[0-9]{2}', text_items[j]['text']):
                        person_info['bank_account'] = text_items[j]['text'].strip()
                        break

    return person_info


def check_signature(text_items):
    """Check if the form has been signed by looking for signature indicators"""
    # Common signature indicators in French and English
    signature_indicators = ['signé', 'signed', 'signature', 'approved', 'approuvé']

    # Look for signature keywords near the bottom of the page
    bottom_items = sorted(text_items, key=lambda x: x['y'], reverse=True)[:20]

    for item in bottom_items:
        text_lower = item['text'].lower()
        if any(indicator in text_lower for indicator in signature_indicators):
            return True

    # Also check for 'required' text that might indicate it needs a signature but doesn't have one
    for item in bottom_items:
        text_lower = item['text'].lower()
        if ('required' in text_lower or 'requis' in text_lower) and not any(
                indicator in text_lower for indicator in signature_indicators):
            return False

    return False  # Default to assuming not signed if no clear indicators


def is_avs_form(text_items):
    """Check if the form is an AVS (travel request) form"""
    form_text = ' '.join([item['text'].lower() for item in text_items])

    # Check for common identifiers of travel request forms
    avs_indicators = [
        'travel request', 'travel authorization', 'demande de voyage',
        'frais de voyage', 'avs', 'mission', 'déplacement'
    ]

    return any(indicator in form_text for indicator in avs_indicators)


def process_pdf_to_excel(pdf_path, excel_path=None):
    """Process a PDF form and extract data to Excel format, targeting April2025 sheet"""
    # Extract text with position information
    text_items = extract_text_from_pdf(pdf_path)

    # Extract personal information
    person_info = extract_person_info(text_items)

    # Extract trip details
    trips = extract_table_data(text_items)

    # Check form metadata
    is_signed = check_signature(text_items)
    is_avs = is_avs_form(text_items)

    # Prepare data for Excel - create one row per trip
    rows = []

    if not trips:
        # If no trip details found, create a single row with person info
        row = {
            'STATUS': 'In Progress',
            'NAME': f"{person_info['first_name']} {person_info['last_name']}".strip(),
            'TRAVEL': '',  # No clear destination
            'DATES': '',  # No clear dates
            'AVS created?': 'yes' if is_avs else 'no',
            'Approved?': 'yes' if is_signed else 'no',
            'SOURCE FILE': os.path.basename(pdf_path)
        }
        rows.append(row)
    else:
        # Create a row for each trip
        for trip in trips:
            row = {
                'STATUS': 'In Progress',
                'NAME': f"{person_info['first_name']} {person_info['last_name']}".strip(),
                'TRAVEL': trip['destination'],
                'DATES': trip['date'],
                'AVS created?': 'yes' if is_avs else 'no',
                'Approved?': 'yes' if is_signed else 'no',
                'SOURCE FILE': os.path.basename(pdf_path)
            }
            rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(rows)

    # Define Excel path if not provided
    if excel_path is None:
        excel_path = 'extracted_form_data.xlsx'

    # Check if file exists
    file_exists = os.path.isfile(excel_path)

    if file_exists:
        # Use ExcelWriter with openpyxl engine to target specific sheet
        try:
            # Load existing workbook
            book = load_workbook(excel_path)

            # Check if April2025 sheet exists
            if 'April2025' not in book.sheetnames:
                print(f"Sheet 'April2025' not found in {excel_path}. Creating new sheet.")
                # If using Pandas to write to a new sheet
                with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
                    df.to_excel(writer, sheet_name='April2025', index=False)
            else:
                # Read existing data from April2025 sheet
                existing_df = pd.read_excel(excel_path, sheet_name='April2025')

                # Append new data
                updated_df = pd.concat([existing_df, df], ignore_index=True)

                # Write back to the same sheet, overwriting previous content
                with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    updated_df.to_excel(writer, sheet_name='April2025', index=False)

            print(f"Data saved to '{excel_path}' in sheet 'April2025'")
        except Exception as e:
            print(f"Error updating Excel file: {e}")
            # Fallback to creating new file if update fails
            df.to_excel(excel_path, sheet_name='April2025', index=False)
    else:
        # Create new Excel file with April2025 sheet
        df.to_excel(excel_path, sheet_name='April2025', index=False)
        print(f"Created new Excel file '{excel_path}' with sheet 'April2025'")

    print(f"Extracted {len(rows)} entries for {person_info['first_name']} {person_info['last_name']}")

    return df


if __name__ == "__main__":
    # Get input from user
    pdf_path = input("Enter the path to the scanned PDF form: ")
    excel_path = input(
        "Enter the path for the output Excel file (leave blank for default): ") or "extracted_form_data.xlsx"

    # Process the form
    df = process_pdf_to_excel(pdf_path, excel_path)

    # Show summary of extracted data
    print("\nExtracted Data Summary:")
    print(df)